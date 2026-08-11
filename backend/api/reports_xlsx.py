"""Server-side Excel report generation for the teacher's per-course
internal marks download. Companion to reports.py (which handles PDFs)."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="4F46E5")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL = PatternFill("solid", fgColor="F3F4F6")


def build_course_internal_marks_xlsx(
    *, institution_name: str, course_info: dict[str, str], rows: list[dict],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Internal Marks"

    ws.merge_cells("A1:N1")
    ws["A1"] = institution_name or "Smart Student Management System"
    ws["A1"].font = Font(bold=True, size=13)

    ws.merge_cells("A2:N2")
    ws["A2"] = "Internal Marks Report"
    ws["A2"].font = Font(bold=True, size=11)

    ws.merge_cells("A3:N3")
    ws["A3"] = f"Generated {datetime.utcnow().strftime('%d %b %Y, %H:%M UTC')}"
    ws["A3"].font = Font(italic=True, size=9, color="6B7280")

    info_row = 5
    for i, (label, value) in enumerate(course_info.items()):
        ws.cell(row=info_row + i, column=1, value=f"{label}:").font = Font(bold=True)
        ws.cell(row=info_row + i, column=2, value=value)

    header_row = info_row + len(course_info) + 2
    headers = [
        "Enrollment", "Name",
        "Att (2)", "Lab (4)", "Exam (8)", "Viva (6)",
        "Att (3)", "Assign (6)", "Pres (3)", "Assess (18)",
        "Practical /20", "Theory /30", "Total /50", "Status",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for i, r in enumerate(rows):
        row_num = header_row + 1 + i
        practical = r["p_att"] + r["p_lab"] + r["p_exam"] + r["p_viva"]
        theory = r["t_att"] + r["t_assign"] + r["t_present"] + r["t_assess"]
        values = [
            r["enrollment"], r["name"],
            r["p_att"], r["p_lab"], r["p_exam"], r["p_viva"],
            r["t_att"], r["t_assign"], r["t_present"], r["t_assess"],
            practical, theory, practical + theory,
            "Published" if r["status"] == "published" else "Draft",
        ]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col, value=v)
            if i % 2 == 1:
                cell.fill = _ALT_FILL

    widths = [14, 24, 8, 8, 8, 8, 8, 9, 8, 10, 11, 10, 10, 11]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()